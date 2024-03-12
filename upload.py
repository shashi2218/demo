from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait
import openpyxl

def update_excel_data(filepath,colName,searchTerm,newValue):
    book = openpyxl.load_workbook(file_path)
    sheet = book.active
    Dict = {}


    for i in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=i).value == searchTerm:
            Dict["col"] = i

    for i in range(1, sheet.max_row + 1):
        for j in range(1, sheet.max_column + 1):
            if sheet.cell(row=i, column=j).value == colName:
                Dict["row"] = i
    sheet.cell(row=Dict["row"], column=Dict["col"]).value = newValue
    book.save(file_path)

file_path ="C:\\Users\\Shashidhar\\Downloads\\download.xlsx"
fruit_name = "Apple"
newValue = "100"
driver = webdriver.Chrome()
driver.implicitly_wait(5)
driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")
driver.find_element(By.ID,"downloadButton").click()
update_excel_data(file_path, fruit_name,"price", newValue)

file_input=driver.find_element(By.CSS_SELECTOR,"input[type='file']")
file_input.send_keys(file_path)

wait =WebDriverWait(driver,5)
toast_locator =(By.CSS_SELECTOR,".Toastify__toast-body div:nth-child(2)")
wait.until(expected_conditions.visibility_of_element_located(toast_locator))
print(driver.find_element(*toast_locator).text)

priceColumn = driver.find_element(By.XPATH,"//div[text()='Price']").get_attribute("data-column-id")
actual_price = driver.find_element(By.XPATH,"//div[text()='"+fruit_name+"']/parent::div/parent::div/div[@id='cell-"+priceColumn+"-undefined']").text
assert actual_price == newValue
print(actual_price)